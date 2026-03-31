"""
JANSA GrandFichier Updater — GrandFichier workbook reader (V1)

Parses all 36 LOT sheets in the GrandFichier workbook into structured GFRow objects.
Handles both layout Variant A (with Zone column) and Variant B (without Zone column).
Reads approbateur groups from Row 8.
"""
import logging
import openpyxl
from pathlib import Path
from typing import Optional

from processing.models import GFRow, GFApprobateur
from processing.config import (
    GF_HEADER_ROW, GF_APPROBATEUR_ROW, GF_SUBHEADER_ROW, GF_DATA_START_ROW,
    GF_COL_VARIANT_A, GF_COL_VARIANT_B,
    GF_APPRO_GROUP_SIZE, GF_OBSERVATIONS_HEADER,
)
from processing.canonical import SAS_REF_PATTERN, normalize_numero

logger = logging.getLogger(__name__)


def _cell_str(ws, row: int, col: int) -> str:
    """Get a cell value as stripped string. col is 1-indexed."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    return str(val).strip()


def _cell_val(ws, row: int, col: int):
    """Get raw cell value. col is 1-indexed."""
    return ws.cell(row=row, column=col).value


def _detect_variant(ws) -> tuple[str, dict]:
    """
    Detect whether a sheet uses Variant A (with Zone column) or Variant B (without).
    Returns ("A", col_map) or ("B", col_map).

    Strategy: inspect Row 7 headers to find Zone column presence.
    """
    # Read headers from row 7
    headers = []
    for col_idx in range(1, 25):
        val = _cell_str(ws, GF_HEADER_ROW, col_idx)
        headers.append(val.upper())

    # Variant A has "ZONE" or "Niv" + "Zone" before N° Doc
    # Check for Zone keyword in first 15 columns
    has_zone = any("ZONE" in h for h in headers[:15])

    if has_zone:
        return "A", GF_COL_VARIANT_A
    else:
        return "B", GF_COL_VARIANT_B


def _read_approbateurs(ws, col_map: dict, max_col: int) -> list[GFApprobateur]:
    """
    Read approbateur groups from Row 8 and Row 9.
    Each approbateur occupies 3 columns: DATE / N° / STATUT.
    Stops when an OBSERVATIONS column is found or row is empty.

    Returns list of GFApprobateur with column positions (0-indexed).
    """
    approbateurs = []
    start_col_0 = col_map["appro_start"]  # 0-indexed
    start_col_1 = start_col_0 + 1         # 1-indexed for openpyxl

    col_1indexed = start_col_1
    while col_1indexed + 2 <= max_col:
        # Read approbateur name from Row 8
        name = _cell_str(ws, GF_APPROBATEUR_ROW, col_1indexed)
        if not name:
            col_1indexed += GF_APPRO_GROUP_SIZE
            continue

        # Stop if we've hit the OBSERVATIONS column
        if GF_OBSERVATIONS_HEADER.upper() in name.upper():
            break

        # Verify sub-headers (Row 9): should be DATE / N° (or similar) / STATUT
        sub1 = _cell_str(ws, GF_SUBHEADER_ROW, col_1indexed).upper()
        sub2 = _cell_str(ws, GF_SUBHEADER_ROW, col_1indexed + 1).upper()
        sub3 = _cell_str(ws, GF_SUBHEADER_ROW, col_1indexed + 2).upper()

        # Accept loosely — different sheets may have slightly different sub-header text
        is_valid_group = (
            ("DATE" in sub1 or "DATE" in sub2) and
            ("STAT" in sub3 or "VISA" in sub3 or "REP" in sub3 or "STATUT" in sub3)
        )
        if not is_valid_group:
            # May still be valid — proceed but log
            logger.debug(
                "Sheet '%s' col %d: approbateur '%s' sub-headers '%s/%s/%s' look unusual",
                ws.title, col_1indexed, name, sub1, sub2, sub3,
            )

        approbateurs.append(GFApprobateur(
            name=name,
            col_date=col_1indexed - 1,       # store 0-indexed
            col_num=col_1indexed,
            col_statut=col_1indexed + 1,
        ))

        col_1indexed += GF_APPRO_GROUP_SIZE

    return approbateurs


def _find_observations_col(ws, approbateurs: list[GFApprobateur], col_map: dict, max_col: int) -> Optional[int]:
    """Find the OBSERVATIONS column index (1-indexed). Returns None if not found."""
    if approbateurs:
        last_appro_end = approbateurs[-1].col_statut + 2  # 0-indexed + 1 = 1-indexed
        # Check columns after last approbateur group
        for c in range(last_appro_end, min(last_appro_end + 4, max_col + 1)):
            val = _cell_str(ws, GF_APPROBATEUR_ROW, c).upper()
            if GF_OBSERVATIONS_HEADER.upper() in val or "OBS" in val:
                return c
    return None


def read_grandfichier(excel_path: str | Path) -> tuple[list[GFRow], dict, dict]:
    """
    Read all LOT sheets from a GrandFichier workbook.

    Returns:
        - list[GFRow]: all data rows across all sheets
        - dict: metadata per sheet (variant, approbateur names, row counts)
        - dict: sas_ref_by_numero — normalized NUMERO → list[GFRow] for rows with SAS REF
    """
    excel_path = Path(excel_path)
    logger.info("Loading GrandFichier: %s", excel_path)

    # Use normal mode (not read_only) so we can also use it for writing later
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)

    all_rows: list[GFRow] = []
    sheet_meta: dict = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_col = ws.max_column or 50
        max_row = ws.max_row or 10

        if max_row < GF_DATA_START_ROW:
            logger.debug("Sheet '%s': fewer than %d rows, skipping", sheet_name, GF_DATA_START_ROW)
            continue

        variant, col_map = _detect_variant(ws)
        approbateurs = _read_approbateurs(ws, col_map, max_col)
        obs_col = _find_observations_col(ws, approbateurs, col_map, max_col)

        sheet_meta[sheet_name] = {
            "variant": variant,
            "approbateur_names": [a.name for a in approbateurs],
            "observations_col": obs_col,
            "row_count": 0,
        }

        appro_col_set = {c for a in approbateurs for c in (a.col_date + 1, a.col_num + 1, a.col_statut + 1)}

        row_count = 0
        for row_num in range(GF_DATA_START_ROW, max_row + 1):
            doc_key = _cell_str(ws, row_num, col_map["document"] + 1)
            if not doc_key:
                continue  # skip blank rows

            titre       = _cell_str(ws, row_num, col_map["titre"] + 1)
            lot         = _cell_str(ws, row_num, col_map["lot"] + 1)
            type_doc    = _cell_str(ws, row_num, col_map["type_doc"] + 1)
            numero      = _cell_str(ws, row_num, col_map["numero"] + 1)
            indice      = _cell_str(ws, row_num, col_map["indice"] + 1)
            visa_global = _cell_str(ws, row_num, col_map["visa_global"] + 1)

            # Niveau and zone vary by variant
            niveau = ""
            zone = ""
            if variant == "A":
                niveau = _cell_str(ws, row_num, col_map.get("niv", col_map.get("niveau", 5)) + 1)
                zone   = _cell_str(ws, row_num, col_map["zone"] + 1)
            else:
                niveau = _cell_str(ws, row_num, col_map.get("niv", col_map.get("niveau", 7)) + 1)

            ancien_val = _cell_val(ws, row_num, col_map["ancien"] + 1)
            ancien = (str(ancien_val).strip() == "1") if ancien_val is not None else False

            # Read observations
            observations = ""
            if obs_col:
                observations = _cell_str(ws, row_num, obs_col)

            # Read date_diffusion (col 2, 0-indexed → col 3, 1-indexed)
            date_diffusion_raw = _cell_val(ws, row_num, col_map["date_diff"] + 1)
            # Keep as raw value — is_same_sas_ref_document handles datetime or str
            date_diffusion_str = ""
            if date_diffusion_raw is not None:
                try:
                    if hasattr(date_diffusion_raw, 'isoformat'):
                        date_diffusion_str = date_diffusion_raw.isoformat()
                    else:
                        date_diffusion_str = str(date_diffusion_raw).strip()
                except Exception:
                    date_diffusion_str = ""

            # V3.1 PATCH 12: Scan ALL cells in this row for SAS REF pattern
            has_sas = False
            for col_scan in range(1, min(max_col + 1, 80)):
                v = ws.cell(row=row_num, column=col_scan).value
                if v is not None and SAS_REF_PATTERN.search(str(v)):
                    has_sas = True
                    break

            # Read current approbateur values
            appro_for_row = []
            for appro in approbateurs:
                cur_date   = _cell_str(ws, row_num, appro.col_date + 1)
                cur_num    = _cell_str(ws, row_num, appro.col_num + 1)
                cur_statut = _cell_str(ws, row_num, appro.col_statut + 1)
                appro_for_row.append(GFApprobateur(
                    name=appro.name,
                    col_date=appro.col_date,
                    col_num=appro.col_num,
                    col_statut=appro.col_statut,
                    current_date=cur_date,
                    current_num=cur_num,
                    current_statut=cur_statut,
                ))

            gf_row = GFRow(
                sheet_name=sheet_name,
                row_number=row_num,
                document_key=doc_key,
                titre=titre,
                lot=lot,
                type_doc=type_doc,
                numero=numero,
                indice=indice,
                niveau=niveau,
                zone=zone,
                ancien=ancien,
                visa_global=visa_global,
                observations=observations,
                approbateurs=appro_for_row,
                has_sas_ref=has_sas,
                date_diffusion=date_diffusion_str,
            )
            all_rows.append(gf_row)
            row_count += 1

        sheet_meta[sheet_name]["row_count"] = row_count
        logger.info("Sheet '%s' (%s): %d rows, %d approbateurs",
                    sheet_name, variant, row_count, len(approbateurs))

    wb.close()

    # V3.1 PATCH 12: Build SAS REF lookup — normalized NUMERO → list of GFRow with SAS REF
    sas_ref_by_numero: dict[str, list[GFRow]] = {}
    sas_ref_count = 0
    for row in all_rows:
        if row.has_sas_ref:
            num = normalize_numero(row.numero)
            if num:
                sas_ref_by_numero.setdefault(num, []).append(row)
                sas_ref_count += 1

    logger.info(
        "GrandFichier read complete: %d total rows across %d sheets, %d rows with SAS REF (%d unique NUMEROs)",
        len(all_rows), len(sheet_meta), sas_ref_count, len(sas_ref_by_numero),
    )
    return all_rows, sheet_meta, sas_ref_by_numero
