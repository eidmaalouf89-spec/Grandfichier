"""
JANSA GrandFichier Updater — GrandFichier workbook writer (V2)

Applies updates to a GrandFichier workbook (in memory) based on DeliverableRecords.
Produces SourceEvidence records for every field written.
Preserves existing formatting.

Update rules (V2):
- VISA GLOBAL: copied from MOEX GEMO STATUT column after approbateur write
- DATE CONTRACTUELLE VISA SYNTHESE: Date réception + 15 calendar days (written once, never overwritten)
- STATUT/DATE per approbateur: resolved via mission_map group → GF row 8 approbateur name
- Never overwrite existing data with empty / pending (PATCH 5)
- OBSERVATIONS: smart update — only new consultant groups added, duplicates skipped
- Groups with no GF column (MOEX SAS, BET VRD, etc.): logged as NO_GF_COLUMN, not written
"""
import csv
import logging
import os
import shutil
import tempfile
from collections import defaultdict
from datetime import timedelta
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from processing.models import DeliverableRecord, CanonicalResponse, SourceEvidence, GFRow, GFApprobateur
from processing.config import GF_DATA_START_ROW, GF_HEADER_ROW
from processing.dates import str_to_date, compare_dates
from processing.anomalies import AnomalyLogger

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# GF standard formatting constants
# Colors extracted from GF header legend cells C3:M5 — DO NOT CHANGE
# ---------------------------------------------------------------------------
GF_FONT = Font(name="Arial Narrow", size=10)
GF_FONT_BOLD = Font(name="Arial Narrow", size=10, bold=True)

# Row fill for ANCIEN rows
GF_ANCIEN_FILL = PatternFill(patternType="solid", fgColor="FFD9D9D9")

# Status color fills — exact hex from the GF header legend
_STATUS_FILLS: dict[str, PatternFill] = {
    "VSO": PatternFill(patternType="solid", fgColor="FFA9D08E"),  # Light green
    "FAV": PatternFill(patternType="solid", fgColor="FFA9D08E"),  # Light green (same as VSO)
    "VAO": PatternFill(patternType="solid", fgColor="FFFFD966"),  # Gold
    "REF": PatternFill(patternType="solid", fgColor="FFF4B083"),  # Salmon/orange
    "DEF": PatternFill(patternType="solid", fgColor="FFFF3300"),  # Red-orange
    "SUS": PatternFill(patternType="solid", fgColor="FFFF3300"),  # Red-orange (same as DEF)
    "HM":  PatternFill(patternType="solid", fgColor="FF9BC2E6"),  # Light blue
    "ANN": PatternFill(patternType="solid", fgColor="FFF2F2F2"),  # Very light grey
}

# Approximate character width for auto row-height calculation (Arial Narrow 10pt)
_OBS_CHARS_PER_LINE = 100  # approximate characters that fit in the OBSERVATIONS column
_OBS_LINE_HEIGHT_PT = 13   # approximate points per line of text


def _get_status_fill(status: str) -> Optional[PatternFill]:
    """Return the fill for a status code, or None if no fill."""
    return _STATUS_FILLS.get(str(status).strip().upper())


def _write_cell_styled(ws, row: int, col: int, value, font=None) -> None:
    """Write a value to a cell with the GF standard font."""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.font = font or GF_FONT


def _write_date_cell(ws, row: int, col: int, date_str: str) -> None:
    """Write a date string as a proper datetime object so Excel recognizes it as a date."""
    from datetime import datetime as dt
    cell = ws.cell(row=row, column=col)
    if not date_str:
        return
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            d = dt.strptime(date_str.strip()[:19], fmt)
            cell.value = d
            cell.font = GF_FONT
            if not cell.number_format or cell.number_format == "General":
                cell.number_format = "DD/MM/YYYY"
            return
        except (ValueError, TypeError):
            continue
    # Fallback: write as string
    cell.value = date_str
    cell.font = GF_FONT


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fmt_date(iso_str: str) -> str:
    """Convert ISO date string to dd/mm/yyyy for GrandFichier display."""
    if not iso_str:
        return ""
    try:
        from datetime import date
        d = date.fromisoformat(iso_str[:10])
        return d.strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return iso_str


def _is_newer(new_iso: str, existing_str: str) -> bool:
    """Return True if new_iso date is newer than existing_str."""
    if not new_iso:
        return False
    if not existing_str:
        return True
    from processing.dates import parse_date
    new_d = str_to_date(new_iso)
    existing_d = parse_date(existing_str)
    return compare_dates(new_d, existing_d) > 0


def should_update(old_value: str, new_value: str, new_status: str = "", is_visa_global: bool = False) -> bool:
    """
    Return True only if the update adds real information.
    - Never write empty or pending statuses
    - For VISA GLOBAL: HM never overwrites a real decision (VAO, VSO, REF, DEF, FAV)
    """
    effective = new_status or new_value
    if not effective:
        return False

    # Never write pending statuses
    if effective in ("EN_ATTENTE", "NONE", ""):
        return False

    # For VISA GLOBAL: HM (Hors Mission) must not overwrite a real visa decision
    if is_visa_global and effective == "HM" and old_value:
        if old_value.upper().strip() in ("VAO", "VSO", "REF", "DEF", "FAV"):
            return False

    return True


# ---------------------------------------------------------------------------
# Mission group resolution helpers (PATCH 3)
# ---------------------------------------------------------------------------

def _get_mission_group(mission: str, mission_map: dict) -> str:
    """Resolve GED mission name → group name via ged_to_group."""
    return mission_map.get("ged_to_group", {}).get(str(mission).strip(), "")


def _resolve_appro_for_group(
    group: str,
    mission_map: dict,
    gf_approbateurs: list[GFApprobateur],
) -> Optional[GFApprobateur]:
    """
    Find the GFApprobateur object for a given mission group on a specific row.
    Uses group_to_gf_appro variants matched case-insensitively against
    the approbateur names already parsed from row 8.
    """
    candidates = mission_map.get("group_to_gf_appro", {}).get(group, [])
    if not candidates:
        return None

    appro_lower = {a.name.lower(): a for a in gf_approbateurs}
    for candidate in candidates:
        c_lower = candidate.lower()
        if c_lower in appro_lower:
            return appro_lower[c_lower]
        # Partial / contains fallback
        for a_lower, appro_obj in appro_lower.items():
            if c_lower in a_lower or a_lower in c_lower:
                return appro_obj
    return None


# Per-sheet column cache — populated once per workbook open (keyed by sheet title)
_visa_global_col_cache: dict[str, int] = {}
_date_contrat_col_cache: dict[str, int] = {}


def _find_visa_global_col(ws, col_map: dict) -> int:
    """
    Find the VISA GLOBAL column (1-indexed) by scanning row 7 for the header text.
    Falls back to col_map["visa_global"] + 1 if not found.
    Cached per sheet to avoid re-scanning on every row.
    """
    key = ws.title
    if key not in _visa_global_col_cache:
        for c in range(1, (ws.max_column or 25) + 1):
            val = str(ws.cell(row=GF_HEADER_ROW, column=c).value or "").replace('\n', ' ').strip().upper()
            if "VISA" in val and "GLOBAL" in val:
                _visa_global_col_cache[key] = c
                break
        else:
            _visa_global_col_cache[key] = col_map.get("visa_global", 14) + 1
    return _visa_global_col_cache[key]


def _find_date_contrat_col(ws, visa_global_col: int) -> int:
    """
    Find the DATE CONTRACTUELLE VISA SYNTHESE column (1-indexed) by scanning row 7.
    Falls back to visa_global_col - 1 if not found.
    Cached per sheet.
    """
    key = ws.title
    if key not in _date_contrat_col_cache:
        for c in range(1, visa_global_col + 1):
            val = str(ws.cell(row=GF_HEADER_ROW, column=c).value or "").replace('\n', ' ').strip().upper()
            if "DATE CONTRACTUELLE" in val or "VISA SYNTHESE" in val:
                _date_contrat_col_cache[key] = c
                break
        else:
            _date_contrat_col_cache[key] = visa_global_col - 1
    return _date_contrat_col_cache[key]


def _compute_date_contractuelle(date_recept_str: str) -> str:
    """
    Compute DATE CONTRACTUELLE VISA SYNTHESE = Date réception + 15 calendar days.
    Input: date string in dd/mm/yyyy or ISO format.
    Output: dd/mm/yyyy string, or "" if unparseable.
    """
    if not date_recept_str:
        return ""
    from processing.dates import parse_date
    d = parse_date(date_recept_str)
    if d is None:
        return ""
    result = d + timedelta(days=15)
    return result.strftime("%d/%m/%Y")


# ---------------------------------------------------------------------------
# Main writer
# ---------------------------------------------------------------------------

def apply_updates(
    source_grandfichier: str | Path,
    deliverables: list[DeliverableRecord],
    gf_rows_by_sheet_row: dict[tuple[str, int], GFRow],
    mission_map: dict,
    source_priority: dict,
    anomaly_logger: AnomalyLogger,
    output_path: str | Path,
    pdf_only: bool = False,
) -> tuple[list[SourceEvidence], int]:
    """
    Open the GrandFichier workbook, apply all updates, and save to output_path.

    Args:
        source_grandfichier: path to the original GrandFichier .xlsx
        deliverables: list of DeliverableRecord (from merge_engine)
        gf_rows_by_sheet_row: dict (sheet_name, row_number) → GFRow
        mission_map: loaded mission_map.json
        source_priority: loaded source_priority.json
        anomaly_logger: AnomalyLogger instance
        output_path: where to save the updated workbook
        pdf_only: if True, only update OBSERVATIONS (PDF pass)

    Returns:
        (evidence_records, fields_updated_count)
    """
    source_grandfichier = Path(source_grandfichier)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    shutil.copy2(str(source_grandfichier), str(output_path))
    logger.info("Copied GrandFichier to: %s", output_path)

    _visa_global_col_cache.clear()
    _date_contrat_col_cache.clear()

    wb = openpyxl.load_workbook(str(output_path))
    evidence: list[SourceEvidence] = []
    fields_updated = 0

    # SAFETY: Record original max row per sheet — writer must NEVER exceed these
    original_max_rows: dict[str, int] = {sn: (wb[sn].max_row or 0) for sn in wb.sheetnames}

    # Track which (sheet, row) pairs were updated, for MAJ column
    edited_rows_per_sheet: dict[str, set[int]] = defaultdict(set)

    # Pre-build mission_map lookups
    no_gf_col_groups: set[str] = set(mission_map.get("no_gf_column", []))
    special_groups: dict[str, str] = mission_map.get("special_groups", {})

    for drec in deliverables:
        sheet_name = drec.gf_sheet
        row_num = drec.gf_row

        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found — skipping row %d", sheet_name, row_num)
            continue

        # SAFETY: Refuse to write beyond the original GF data range
        assert row_num <= original_max_rows.get(sheet_name, 0), (
            f"BLOCKED: Attempted to write to row {row_num} in '{sheet_name}' "
            f"(original max={original_max_rows.get(sheet_name, 0)})"
        )

        ws = wb[sheet_name]
        gf_key = (sheet_name, row_num)
        gf_row = gf_rows_by_sheet_row.get(gf_key)

        if gf_row is None:
            logger.warning("GFRow not found for (%s, %d)", sheet_name, row_num)
            continue

        from processing.grandfichier_reader import _detect_variant
        _, col_map = _detect_variant(ws)

        is_ancien = gf_row.ancien

        # ── Group responses by mission (deduplicate: keep best per mission×field) ──
        by_mission: dict[str, list[CanonicalResponse]] = defaultdict(list)
        for cr in drec.responses:
            by_mission[cr.mission].append(cr)

        priority_status = source_priority.get("status", ["GED", "SAS", "REPORT"])
        priority_date   = source_priority.get("response_date", ["GED", "SAS", "REPORT"])

        for mission, responses in by_mission.items():
            group = _get_mission_group(mission, mission_map)

            # ── Skip (MOEX SAS): do nothing ──
            if special_groups.get(group) == "SKIP":
                continue

            # ── No GF column: log anomaly and skip ──
            if group in no_gf_col_groups:
                for cr in responses:
                    anomaly_logger.log_no_gf_column(
                        source_file=cr.source_file,
                        source_row=cr.source_row_or_page,
                        document_key=cr.document_key,
                        mission_name=mission,
                        group_name=group,
                        status=cr.normalized_status,
                        response_date=cr.response_date,
                        comment=cr.comment,
                    )
                continue

            # ── PDF-only pass: only OBSERVATIONS ──
            if pdf_only:
                _append_observations_from_responses(
                    ws, gf_row, responses, row_num, sheet_name, evidence, mission_map
                )
                continue

            # ── Find approbateur columns for this group ──
            appro = _resolve_appro_for_group(group, mission_map, gf_row.approbateurs)
            if appro is None:
                if group:
                    logger.debug(
                        "Sheet '%s' row %d: group '%s' (mission '%s') — no matching approbateur column",
                        sheet_name, row_num, group, mission,
                    )
                continue

            # Pick best response for status and date
            best_status = _pick_best_by_priority(responses, "normalized_status", priority_status)
            best_date   = _pick_best_by_priority(responses, "response_date", priority_date)

            # ── Write STATUT ──
            if best_status and best_status.normalized_status:
                new_s = best_status.normalized_status
                old_s = appro.current_statut
                if should_update(old_s, new_s, new_s) and new_s != old_s:
                    _write_cell_styled(ws, row_num, appro.col_statut + 1, new_s)
                    ev = SourceEvidence(
                        sheet_name=sheet_name,
                        row_number=row_num,
                        column_name=f"STATUT ({appro.name})",
                        old_value=old_s,
                        new_value=new_s,
                        source_type=best_status.source_type,
                        source_file=best_status.source_file,
                        source_row_or_page=best_status.source_row_or_page,
                        update_reason=f"GED group '{group}' → appro '{appro.name}' ({best_status.match_strategy})",
                    )
                    if is_ancien:
                        ev.update_reason += " [ANCIEN row]"
                    evidence.append(ev)
                    fields_updated += 1
                    edited_rows_per_sheet[sheet_name].add(row_num)

            # ── Write DATE ──
            if best_date and best_date.response_date:
                old_d = appro.current_date
                if should_update(old_d, best_date.response_date) and _is_newer(best_date.response_date, old_d):
                    new_d_display = _fmt_date(best_date.response_date)
                    _write_date_cell(ws, row_num, appro.col_date + 1, new_d_display)
                    evidence.append(SourceEvidence(
                        sheet_name=sheet_name,
                        row_number=row_num,
                        column_name=f"DATE ({appro.name})",
                        old_value=old_d,
                        new_value=new_d_display,
                        source_type=best_date.source_type,
                        source_file=best_date.source_file,
                        source_row_or_page=best_date.source_row_or_page,
                        update_reason="GED response date is newer",
                    ))
                    fields_updated += 1
                    edited_rows_per_sheet[sheet_name].add(row_num)

        # ── VISA GLOBAL = copy of MOEX GEMO STATUT ──
        # ── DATE CONTRACTUELLE VISA SYNTHESE = Date réception + 15 jours ──
        if not pdf_only:
            visa_col = _find_visa_global_col(ws, col_map)
            date_cont_col = _find_date_contrat_col(ws, visa_col)

            # Find the MOEX GEMO approbateur object for this row
            moex_appro = None
            for a in gf_row.approbateurs:
                if "MOEX" in a.name.upper() or "GEMO" in a.name.upper():
                    moex_appro = a
                    break

            # Step 1: Copy MOEX GEMO STATUT → VISA GLOBAL
            if moex_appro is not None:
                moex_statut = str(ws.cell(row=row_num, column=moex_appro.col_statut + 1).value or "")
                old_vg = str(ws.cell(row=row_num, column=visa_col).value or "")
                if moex_statut and should_update(old_vg, moex_statut, moex_statut, is_visa_global=True) and moex_statut != old_vg:
                    _write_cell_styled(ws, row_num, visa_col, moex_statut)
                    evidence.append(SourceEvidence(
                        sheet_name=sheet_name,
                        row_number=row_num,
                        column_name="VISA GLOBAL",
                        old_value=old_vg,
                        new_value=moex_statut,
                        source_type="GF_COPY",
                        source_file="",
                        source_row_or_page="",
                        update_reason="Copied from MOEX GEMO STATUT column",
                    ))
                    fields_updated += 1
                    edited_rows_per_sheet[sheet_name].add(row_num)

            # Step 2: DATE CONTRACTUELLE = Date réception + 15 jours
            old_dc = str(ws.cell(row=row_num, column=date_cont_col).value or "")
            if not old_dc.strip():
                date_recept = str(gf_row.date_recept or "")
                new_dc = _compute_date_contractuelle(date_recept)
                if new_dc:
                    _write_date_cell(ws, row_num, date_cont_col, new_dc)
                    evidence.append(SourceEvidence(
                        sheet_name=sheet_name,
                        row_number=row_num,
                        column_name="DATE CONTRACTUELLE VISA SYNTHESE",
                        old_value=old_dc,
                        new_value=new_dc,
                        source_type="COMPUTED",
                        source_file="",
                        source_row_or_page="",
                        update_reason="Date réception + 15 calendar days",
                    ))
                    fields_updated += 1
                    edited_rows_per_sheet[sheet_name].add(row_num)

        # ── Append OBSERVATIONS (GED pass only — PDF appends handled separately) ──
        if not pdf_only:
            _append_observations_from_responses(
                ws, gf_row, drec.responses, row_num, sheet_name, evidence, mission_map
            )

    # ── Write MAJ header + "Edited" label for every updated row ──
    if edited_rows_per_sheet:
        for sname, row_set in edited_rows_per_sheet.items():
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            # Find OBSERVATIONS column in row 8
            obs_col = None
            for c in range(1, (ws.max_column or 60) + 1):
                val = str(ws.cell(row=8, column=c).value or "").strip().upper()
                if val == "OBSERVATIONS":
                    obs_col = c
                    break
            if not obs_col:
                continue
            edited_col = obs_col + 1
            # Write MAJ header in row 7 if not already present
            if not ws.cell(row=7, column=edited_col).value:
                _write_cell_styled(ws, 7, edited_col, "MAJ", font=GF_FONT_BOLD)
            # Write "Edited" for each updated row
            for row_num in row_set:
                _write_cell_styled(ws, row_num, edited_col, "Edited")
        logger.info(
            "Wrote 'Edited' label for %d rows across %d sheets",
            sum(len(v) for v in edited_rows_per_sheet.values()),
            len(edited_rows_per_sheet),
        )

    # Format and cleanup: row coloring, filters, print setup, trim blank rows
    _format_and_cleanup(wb)

    # Save via /tmp to avoid corrupting FUSE state (writing a large xlsx directly
    # to the FUSE-mounted filesystem breaks subsequent open() calls in this process).
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as _tmp:
        _tmp_xlsx = _tmp.name
    wb.save(_tmp_xlsx)
    shutil.copy2(_tmp_xlsx, str(output_path))
    os.remove(_tmp_xlsx)
    logger.info(
        "Writer complete: %d fields updated, %d evidence records, saved to %s",
        fields_updated, len(evidence), output_path,
    )
    return evidence, fields_updated


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _pick_best_by_priority(
    responses: list[CanonicalResponse],
    attr: str,
    priority_list: list[str],
) -> Optional[CanonicalResponse]:
    """Pick the response with the highest-priority source that has a non-empty value for attr."""
    candidates = [r for r in responses if getattr(r, attr, "")]
    if not candidates:
        return None
    return min(candidates, key=lambda r: _source_rank(r.source_type, priority_list))


def _source_rank(source_type: str, priority_list: list[str]) -> int:
    try:
        return priority_list.index(source_type)
    except ValueError:
        return len(priority_list) + 99


def _get_col_map(ws, gf_row: GFRow) -> Optional[dict]:
    from processing.grandfichier_reader import _detect_variant
    try:
        _, col_map = _detect_variant(ws)
        return col_map
    except Exception:
        return None


# Phrases that indicate "no real comment — placeholder only".
# These use startswith matching: if the comment begins with one of these phrases, it's a placeholder.
# Only multi-word phrases go here — never single characters or very short tokens.
_EMPTY_COMMENT_PREFIX_PATTERNS = [
    "voir documents joints",
    "voir document joint",
    "voir doc joint",
    "voir pièce jointe",
    "voir pièces jointes",
    "voir pièces joints",
    "voir pj",
    "voir visa",
    "voir note",
    "voir annotation",
    "voir fichier joint",
    "voir fichier",
    "sans observation",
    "pas d'observation",
    "pas de remarque",
    "aucune observation",
    "aucune remarque",
    "non concerné",
    "hors mission",
    "document non visé",
    "non visé",
]

# Exact-match-only patterns: the ENTIRE comment (stripped, lowercased) must equal one of these.
# Short placeholders that should NOT use startswith (to avoid killing real comments
# starting with "-", ".", "ok" etc.)
_EMPTY_COMMENT_EXACT_PATTERNS = {
    "-", ".", "..", "...", "ok", "ok.", "n/a", "na", "/",
    "ras", "r.a.s", "r.a.s.", "rsa", "r a s",
    "néant", "neant",
    "none", "rien",
    "sans obs", "sans obs.",
    "x",
}


def _is_empty_comment(comment: str) -> bool:
    """
    Return True if the comment has no real content (placeholder or empty).

    Strategy:
    1. Empty or whitespace-only → empty
    2. Very short (< 3 chars after stripping) → empty
    3. Exact match against short placeholder tokens (e.g., "-", "ok", "ras") → empty
    4. Starts with a known placeholder phrase (e.g., "voir documents joints") → empty
    5. Everything else → real comment, keep it
    """
    if not comment:
        return True
    clean = comment.strip().lower()
    if len(clean) < 3:
        return True
    # Exact match for short placeholders — the WHOLE comment must be just this token
    if clean in _EMPTY_COMMENT_EXACT_PATTERNS:
        return True
    # Prefix match for multi-word placeholder phrases only
    if any(clean.startswith(p) for p in _EMPTY_COMMENT_PREFIX_PATTERNS):
        return True
    return False


def _extract_existing_groups(observations_text: str) -> set:
    """
    Parse existing OBSERVATIONS text to find which consultant groups
    already have responses recorded.

    Detects patterns like:
    - "GEMO : VAO"
    - "BET STR:VSO"
    - "ARCHI : REF"
    - "ARCHI MOX : VAO"
    - "ACOUSTICIEN : HM"

    Returns a set of normalized group name prefixes found.
    """
    if not observations_text:
        return set()

    import re
    found = set()

    pattern = re.compile(
        r'([A-ZÀ-Ÿ][A-ZÀ-Ÿ\s\-\.]{2,30}?)\s*:\s*'
        r'(VAO|VSO|REF|DEF|HM|FAV|SUS|SUSPENDU|EN.ATTENTE|FAVORABLE|DEFAVORABLE)',
        re.IGNORECASE
    )

    for match in pattern.finditer(observations_text.upper()):
        group_name = match.group(1).strip()
        normalized = _normalize_obs_group(group_name)
        if normalized:
            found.add(normalized)

    return found


# Comprehensive mapping: OBSERVATIONS group name → unified mission_map group
# Built from scanning all 3,249 unique group-name patterns across production GF.
# Case-insensitive lookup is done by the caller (.upper() on key before lookup).
_OBS_GROUP_NORMALIZE_MAP: dict[str, str] = {
    # ── MOEX — always written as GEMO in OBSERVATIONS, never as "MOX" ──
    'GEMO': 'MOEX',
    'MOEX': 'MOEX',
    'MOEX GEMO': 'MOEX',
    'VISA GEMO': 'MOEX',
    'GEMO / MOA': 'MOEX',
    'GEMO/MOA': 'MOEX',

    # ── MOEX SAS variants — SKIP group ──
    'SAS': 'MOEX SAS',
    'GEMO SAS': 'MOEX SAS',
    'GEMO:SAS': 'MOEX SAS',
    'GEMO: SAS': 'MOEX SAS',
    'GEMO-SAS': 'MOEX SAS',
    'GEMO  SAS': 'MOEX SAS',
    'GEMO :  SAS': 'MOEX SAS',
    'GEMO : SAS': 'MOEX SAS',
    'GEMO: SASA': 'MOEX SAS',
    'GEMO SAS REF': 'MOEX SAS',
    'GEMO SA': 'MOEX SAS',
    'GEMO: MOX': 'MOEX SAS',

    # ── ARCHITECTE — MOX = ARCHI MOX shorthand, NOT MOEX ──
    'MOX': 'ARCHITECTE',
    'ARCHI': 'ARCHITECTE',
    'ARCHI MOX': 'ARCHITECTE',
    'ARCHITECTE': 'ARCHITECTE',
    'ARCHITECTE MOX': 'ARCHITECTE',
    'ARCHITECTES': 'ARCHITECTE',
    'MOX ARCHI': 'ARCHITECTE',
    'ARCHIMOX': 'ARCHITECTE',
    'ARCHBI MOX': 'ARCHITECTE',
    'ARCHII': 'ARCHITECTE',
    'ARCHII MOX': 'ARCHITECTE',
    'ARCHIO': 'ARCHITECTE',
    'ARCHIA': 'ARCHITECTE',
    'ARCI': 'ARCHITECTE',
    'ARCI MOX': 'ARCHITECTE',
    'ACHI': 'ARCHITECTE',
    'ARCH MOX': 'ARCHITECTE',
    'ARECHI MOX': 'ARCHITECTE',
    'ARHITECTE': 'ARCHITECTE',
    'ARCHITEECTE': 'ARCHITECTE',
    'ARCHITCTES': 'ARCHITECTE',
    'B-ARCHITECTE': 'ARCHITECTE',
    'B- MOX': 'ARCHITECTE',
    'ARCHIO MOX': 'ARCHITECTE',
    'ARCHBI': 'ARCHITECTE',
    'ARCHI /': 'ARCHITECTE',
    'ARCH /': 'ARCHITECTE',

    # ── BET Structure ──
    'BET STR': 'BET Structure',
    'BET STR-TERRELL': 'BET Structure',
    'BET STR TERRELL': 'BET Structure',
    'STR-TERRELL': 'BET Structure',
    'TERRELL': 'BET Structure',
    'TERREL': 'BET Structure',
    'TERELLE': 'BET Structure',
    'TERELL': 'BET Structure',
    'BET TERRELL': 'BET Structure',
    'BET STRUCTURE': 'BET Structure',
    'BET STRUCTURE TERRELL': 'BET Structure',
    'BET TRL': 'BET Structure',
    'BET TER': 'BET Structure',
    'BET SRT': 'BET Structure',
    'BET STRUCRURE': 'BET Structure',
    'BET STRE TERRELL': 'BET Structure',
    'BET TSR TERRELL': 'BET Structure',
    'BET STR TERRRELLL': 'BET Structure',
    'BETSTR': 'BET Structure',
    'BUT STRUCTURE': 'BET Structure',
    'TERRELL STR': 'BET Structure',
    'TERREL STR': 'BET Structure',
    'TERRELL:STR': 'BET Structure',
    'B-BET STRUCTURE TERRELL': 'BET Structure',
    'BET STR TERRELL /': 'BET Structure',

    # ── Bureau de contrôle ──
    'SOCOTEC': 'Bureau de control',
    'BC SOCOTEC': 'Bureau de control',
    'BC': 'Bureau de control',
    'CT SOCOTEC': 'Bureau de control',
    'BC SOSCOTEC': 'Bureau de control',
    'BUREAU DE CONTRÔLE': 'Bureau de control',
    'BUREAU DE CONTROLE': 'Bureau de control',
    'BET CONTROLE': 'Bureau de control',

    # ── AMO HQE ──
    'AMO HQE': 'AMO HQE',
    'AMO HQE LE SOMMER': 'AMO HQE',
    'LE SOMMER': 'AMO HQE',
    'HQE': 'AMO HQE',
    'AMO': 'AMO HQE',
    'AMO ENV LE SOMMER': 'AMO HQE',
    'AMO ENV LESOMMER': 'AMO HQE',
    'B-AMO HQE': 'AMO HQE',
    '-AMO HQE': 'AMO HQE',
    'AMO: HQE': 'AMO HQE',

    # ── BET Géotech ──
    'GEOLIA': 'BET Géotech',
    'BET GEOLIA': 'BET Géotech',
    'BET GEOLIA - G4': 'BET Géotech',
    'BET GEOTECH GEOLIA': 'BET Géotech',
    'G4': 'BET Géotech',

    # ── BET ACOUSTIQUE ──
    'ACOUSTICIEN': 'BET ACOUSTIQUE',
    'ACOUSTICIEN AVLS': 'BET ACOUSTIQUE',
    'AVLS': 'BET ACOUSTIQUE',
    'BET AVLS': 'BET ACOUSTIQUE',
    'BET ACOUSTIQUE': 'BET ACOUSTIQUE',
    'BET ACOUS AVLS': 'BET ACOUSTIQUE',
    'BET ACOUST AVLS': 'BET ACOUSTIQUE',
    'ACOUS': 'BET ACOUSTIQUE',
    'BET ACOUSTIQUE AVLS': 'BET ACOUSTIQUE',

    # ── BET POL ──
    'BET POLLUTION': 'BET POL',
    'BET POLLUTION DIE': 'BET POL',
    'POLLUTION DIE': 'BET POL',
    'DIE': 'BET POL',
    'BET DIE': 'BET POL',
    'BET POL': 'BET POL',
    'BET POL DIE': 'BET POL',
    'DIE POLLUTION': 'BET POL',

    # ── BET CVC ──
    'BET CVC': 'BET CVC',
    'BET CVC - EGIS': 'BET CVC',
    'BET CVC EGIS': 'BET CVC',
    'BET EGIS CVC': 'BET CVC',
    '-BET CVC': 'BET CVC',

    # ── BET Plomberie ──
    'BET PLOMB': 'BET Plomberie',
    'BET PLOMB - EGIS': 'BET Plomberie',
    'BET PLOMBERIE': 'BET Plomberie',
    'BET PLOMBERIE EGIS': 'BET Plomberie',
    'BET PLMB EGIS': 'BET Plomberie',
    'BET PLOMBERIE  EGIS': 'BET Plomberie',
    'BET PLB': 'BET Plomberie',
    'BET PLOMBERIE ERGIS': 'BET Plomberie',

    # ── BET ELEC ──
    'BET EGIS': 'BET ELEC',
    'EGIS': 'BET ELEC',
    'BET ELEC': 'BET ELEC',
    'BET ELECTRICITÉ': 'BET ELEC',
    'BET ELECTRICITE': 'BET ELEC',
    'BET ELECTRICITÉ EGIS': 'BET ELEC',
    'BET ELECTRICITE EGIS': 'BET ELEC',
    'BET ELEC EGIS': 'BET ELEC',
    'BET ELGIS': 'BET ELEC',
    'BET GEIS': 'BET ELEC',
    'EIGS': 'BET ELEC',
    'EGI': 'BET ELEC',

    # ── BET Façade ──
    'BET FACADE': 'BET Façade',
    'BET FACADE - ELIOTH': 'BET Façade',
    'BET FACADE ELIOTH': 'BET Façade',
    'BET ELIOTH': 'BET Façade',
    'ELIOTH': 'BET Façade',
    'BET FAÇADE': 'BET Façade',
    'BET FAÇADE ELIOTH': 'BET Façade',
    'BET FAC ELIOTH': 'BET Façade',
    'BET FAC': 'BET Façade',
    'BET FACADES': 'BET Façade',
    'ELLIOTH': 'BET Façade',
    '-BET FAÇADE': 'BET Façade',
    '-BET FACADE': 'BET Façade',

    # ── BET SPK ──
    'BET SPK': 'BET SPK',
    'SPK': 'BET SPK',

    # ── BET Ascenseur ──
    'BET ASCENSEUR': 'BET Ascenseur',
    'BET ASCAUDIT': 'BET Ascenseur',
    'ASCAUDIT': 'BET Ascenseur',
    'BET ASC': 'BET Ascenseur',

    # ── BET EV ──
    'MUGO': 'BET EV',
    'PAYSAGISTE MUGO': 'BET EV',
    'MUGO PAYSAGISTE': 'BET EV',
}


def _normalize_obs_group(raw_name: str) -> str:
    """
    Normalize a group name found in OBSERVATIONS to match the unified
    group names from mission_map.json.

    Strategy:
    1. Exact match against comprehensive lookup table (case-insensitive)
    2. Partial / contains fallback for remaining variants
    """
    name = raw_name.strip().upper()

    # 1. Exact match (fast path — covers 95%+ of cases)
    if name in _OBS_GROUP_NORMALIZE_MAP:
        return _OBS_GROUP_NORMALIZE_MAP[name]

    # 2. Partial / contains fallback for typos and rare variants
    # Check longest keys first to avoid short-key false positives
    for key in sorted(_OBS_GROUP_NORMALIZE_MAP.keys(), key=len, reverse=True):
        if key in name or name in key:
            return _OBS_GROUP_NORMALIZE_MAP[key]

    return name


def _build_obs_entry(group_display_name: str, status: str, comment: str) -> str:
    """
    Build a single OBSERVATIONS entry in the standard format.

    Format:
        GROUP_NAME : STATUS
        comment text (if not empty/placeholder)
    """
    header = f"{group_display_name} : {status}"

    if _is_empty_comment(comment):
        return header

    return f"{header}\n{comment}"


# Display name mapping: unified group → short display name for OBSERVATIONS
_GROUP_DISPLAY_NAMES = {
    'MOEX': 'GEMO',
    'ARCHITECTE': 'ARCHI MOX',
    'BET Structure': 'BET STR',
    'Bureau de control': 'SOCOTEC',
    'AMO HQE': 'AMO HQE',
    'BET Géotech': 'GEOLIA',
    'BET ACOUSTIQUE': 'ACOUSTICIEN',
    'BET POL': 'BET POLLUTION',
    'BET CVC': 'BET CVC',
    'BET Plomberie': 'BET PLOMB',
    'BET ELEC': 'BET EGIS',
    'BET Façade': 'BET FACADE',
    'BET Ascenseur': 'BET ASCENSEUR',
    'BET SPK': 'BET SPK',
    'BET EV': 'PAYSAGISTE MUGO',
}


def _append_observations_from_responses(
    ws,
    gf_row: GFRow,
    responses: list[CanonicalResponse],
    row_num: int,
    sheet_name: str,
    evidence: list[SourceEvidence],
    mission_map: dict,
) -> None:
    """
    Smart OBSERVATIONS update:
    1. Read existing cell content
    2. Detect which groups already have responses
    3. Only add NEW responses not already present
    4. Format: GROUP_NAME : STATUS + comment
    5. Skip placeholder comments ("voir documents joints" etc.)
    """
    from processing.grandfichier_reader import _find_observations_col, _detect_variant, _read_approbateurs

    _, col_map = _detect_variant(ws)
    appros = gf_row.approbateurs
    obs_col = _find_observations_col(ws, appros, col_map, ws.max_column or 60)
    if not obs_col:
        return

    existing_obs = str(ws.cell(row=row_num, column=obs_col).value or "")

    # Parse which groups already have responses in existing text
    existing_groups = _extract_existing_groups(existing_obs)

    # Group responses by mission group
    ged_to_group = mission_map.get("ged_to_group", {})
    special_groups = mission_map.get("special_groups", {})
    no_gf_col = set(mission_map.get("no_gf_column", []))

    new_entries = []
    for cr in responses:
        if not cr.comment and not cr.normalized_status:
            continue

        group = ged_to_group.get(cr.mission, "")
        if not group:
            continue

        # Skip MOEX SAS and groups without GF column
        if special_groups.get(group) == "SKIP":
            continue
        if group in no_gf_col:
            continue

        # Skip if this group already has a response in existing OBSERVATIONS
        if group in existing_groups:
            continue

        # Skip if no real status
        if not cr.normalized_status or cr.normalized_status in ("EN_ATTENTE", "NONE", ""):
            continue

        # Build display name
        display_name = _GROUP_DISPLAY_NAMES.get(group, group)

        # Build the entry
        entry = _build_obs_entry(display_name, cr.normalized_status, cr.comment)
        new_entries.append((group, entry))

    # Deduplicate: one entry per group (take the first/best)
    seen_groups: set = set()
    final_entries = []
    for group, entry in new_entries:
        if group not in seen_groups:
            seen_groups.add(group)
            final_entries.append(entry)

    if not final_entries:
        return

    # Append to existing observations
    separator = "\n" if existing_obs.strip() else ""
    new_text = "\n".join(final_entries)
    updated_obs = existing_obs + separator + new_text if existing_obs.strip() else new_text

    cell = ws.cell(row=row_num, column=obs_col)
    cell.value = updated_obs
    cell.font = GF_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")

    evidence.append(SourceEvidence(
        sheet_name=sheet_name,
        row_number=row_num,
        column_name="OBSERVATIONS",
        old_value=existing_obs[:80] + "..." if len(existing_obs) > 80 else existing_obs,
        new_value=f"[added {len(final_entries)} new response(s): {', '.join(g for g, _ in new_entries[:3])}]",
        source_type="GED",
        source_file="",
        source_row_or_page="",
        update_reason="Smart OBS update: only new responses added, duplicates skipped",
    ))


# ---------------------------------------------------------------------------
# Post-processing: formatting, print setup, filters
# ---------------------------------------------------------------------------

def _format_and_cleanup(wb) -> None:
    """
    Post-processing pass on the entire output workbook:
    1. Row coloring: ancien → grey, visa global filled → row color, else → white with per-cell status colors
    2. OBSERVATIONS: auto row height based on content
    3. AutoFilter on row 9
    4. Print setup: area, rows to repeat, fit to 1 page wide, landscape
    5. Trim trailing blank rows
    """
    from processing.config import GF_DATA_START_ROW, GF_HEADER_ROW

    for sn in wb.sheetnames:
        ws = wb[sn]
        max_col = ws.max_column or 10
        if (ws.max_row or 0) < GF_DATA_START_ROW:
            continue

        # --- Detect key columns ---
        # Find VISA GLOBAL col (1-indexed)
        visa_col = None
        for c in range(1, min(max_col + 1, 25)):
            val = str(ws.cell(row=GF_HEADER_ROW, column=c).value or "").replace('\n', ' ').strip().upper()
            if "VISA" in val and "GLOBAL" in val:
                visa_col = c
                break

        # Find ANCIEN col (1-indexed)
        ancien_col = None
        for c in range(1, min(max_col + 1, 20)):
            val = str(ws.cell(row=GF_HEADER_ROW, column=c).value or "").replace('\n', ' ').strip().upper()
            if "ANCIEN" in val:
                ancien_col = c
                break

        # Find OBSERVATIONS col (1-indexed) — from row 8
        obs_col = None
        for c in range(1, max_col + 1):
            val = str(ws.cell(row=8, column=c).value or "").strip().upper()
            if val == "OBSERVATIONS":
                obs_col = c
                break

        # Find all STATUT columns (1-indexed) — from row 9 sub-headers
        statut_cols = []
        for c in range(1, max_col + 1):
            val = str(ws.cell(row=9, column=c).value or "").strip().upper()
            if "STAT" in val or val == "STATUT":
                statut_cols.append(c)

        # --- Find last data row ---
        last_data_row = GF_DATA_START_ROW - 1
        for r in range(ws.max_row, GF_DATA_START_ROW - 1, -1):
            if any(ws.cell(row=r, column=c).value is not None for c in range(1, min(max_col, 10))):
                last_data_row = r
                break

        # --- Row coloring + OBSERVATIONS auto height ---
        for r in range(GF_DATA_START_ROW, last_data_row + 1):
            # Check if row has data
            doc_val = ws.cell(row=r, column=1).value
            if doc_val is None:
                continue

            # Check ANCIEN flag
            is_ancien = False
            if ancien_col:
                anc_val = ws.cell(row=r, column=ancien_col).value
                is_ancien = (str(anc_val).strip() == "1") if anc_val is not None else False

            # Check VISA GLOBAL value
            visa_val = ""
            if visa_col:
                visa_val = str(ws.cell(row=r, column=visa_col).value or "").strip().upper()

            # Determine the fill to apply to the row
            last_col_for_fill = obs_col or max_col  # fill up to OBSERVATIONS col

            if is_ancien:
                # Rule 1 (highest priority, overrides all others):
                # ANCIEN = 1 → grey out the ENTIRE row, including all columns beyond OBSERVATIONS
                for c in range(1, max_col + 1):
                    ws.cell(row=r, column=c).fill = GF_ANCIEN_FILL

            elif visa_val and visa_val not in ("", "EN_ATTENTE", "NONE"):
                # Rule 2: VISA GLOBAL filled (and ANCIEN != 1) → entire row takes visa color
                # e.g. VAO → #FFD966 across ALL columns; overrides per-cell status colors
                row_fill = _get_status_fill(visa_val)
                if row_fill:
                    for c in range(1, max_col + 1):
                        ws.cell(row=r, column=c).fill = row_fill

            else:
                # Rule 3: No VISA GLOBAL → white row, per-cell STATUT color coding only
                for sc in statut_cols:
                    stat_val = str(ws.cell(row=r, column=sc).value or "").strip().upper()
                    if stat_val:
                        sfill = _get_status_fill(stat_val)
                        if sfill:
                            ws.cell(row=r, column=sc).fill = sfill

            # --- OBSERVATIONS: auto row height ---
            if obs_col:
                obs_cell = ws.cell(row=r, column=obs_col)
                obs_text = str(obs_cell.value or "")
                obs_cell.alignment = Alignment(wrap_text=True, vertical="top")
                if obs_text:
                    # Estimate required height
                    line_count = obs_text.count('\n') + 1
                    # Also account for long lines wrapping
                    for line in obs_text.split('\n'):
                        if len(line) > _OBS_CHARS_PER_LINE:
                            line_count += len(line) // _OBS_CHARS_PER_LINE
                    estimated_height = max(line_count * _OBS_LINE_HEIGHT_PT, 15)
                    # Only increase height, never shrink below current
                    current_height = ws.row_dimensions[r].height or 15
                    if estimated_height > current_height:
                        ws.row_dimensions[r].height = estimated_height

        # --- AutoFilter on row 9 ---
        filter_end_col = obs_col or max_col
        filter_range = f"A9:{get_column_letter(filter_end_col)}9"
        ws.auto_filter.ref = filter_range

        # --- Print setup ---
        # Print area: col A to OBSERVATIONS column, row 1 to last data row
        if obs_col:
            print_area = f"A1:{get_column_letter(obs_col)}{last_data_row}"
            ws.print_area = print_area

        # Rows to repeat at top: 1 through 9
        ws.print_title_rows = "1:9"

        # Fit to 1 page wide, auto height
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 0 = auto (as many pages as needed vertically)
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # Landscape orientation
        ws.page_setup.orientation = "landscape"

        # --- Trim trailing blank rows ---
        keep_until = last_data_row + 5
        if ws.max_row > keep_until:
            ws.delete_rows(keep_until + 1, ws.max_row - keep_until)


# ---------------------------------------------------------------------------
# Orphan GED export
# ---------------------------------------------------------------------------

def export_orphan_ged(
    orphan_records: list[CanonicalResponse],
    all_ged_records: list[CanonicalResponse],
    output_path: Path,
) -> None:
    """
    Export orphan GED documents where MOEX is still En attente.

    Only includes docs where:
      1. The document is NOT in the GF (orphan — caller guarantees this)
      2. The document has at least one MOEX mission response
      3. ALL MOEX responses are "En attente" (open — MOEX has not yet responded)

    Deduplicates by (NUMERO, INDICE), sorts by EMETTEUR then NUMERO.
    TITRE column shows the actual Libellé du fichier from the GED.
    Saves via /tmp to avoid FUSE write issues.
    """
    from collections import defaultdict
    from processing.actors import MOEX_MISSIONS

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Build (NUMERO, INDICE) → all responses, to check MOEX status
    doc_responses: dict[tuple, list[CanonicalResponse]] = defaultdict(list)
    for cr in all_ged_records:
        doc_responses[(cr.numero, cr.indice)].append(cr)

    # Filter orphan records to only those where MOEX is En attente
    filtered: list[CanonicalResponse] = []
    seen: set[tuple] = set()
    for cr in orphan_records:
        key = (cr.numero or "", cr.indice or "")
        if key in seen:
            continue

        responses = doc_responses.get((cr.numero, cr.indice), [])

        # Must have at least one MOEX mission
        moex_responses = [r for r in responses if r.mission in MOEX_MISSIONS]
        if not moex_responses:
            continue

        # All MOEX responses must be En attente
        moex_statuses = [r.raw_status.strip().lower() for r in moex_responses if r.raw_status]
        if not moex_statuses or not all(s == "en attente" for s in moex_statuses):
            continue  # MOEX already responded → disregard

        seen.add(key)
        filtered.append(cr)

    # Sort: EMETTEUR then NUMERO
    filtered.sort(key=lambda r: (r.emetteur or "", r.numero or ""))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orphan GED"

    headers = ["EMETTEUR", "LOT", "TYPE_DOC", "NUMERO", "INDICE", "TITRE", "DATE", "MISSIONS"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = h

    for row_idx, cr in enumerate(filtered, start=2):
        ws.cell(row=row_idx, column=1).value = cr.emetteur or ""
        ws.cell(row=row_idx, column=2).value = cr.lot or ""
        ws.cell(row=row_idx, column=3).value = cr.type_doc or ""
        ws.cell(row=row_idx, column=4).value = cr.numero or ""
        ws.cell(row=row_idx, column=5).value = cr.indice or ""
        ws.cell(row=row_idx, column=6).value = cr.libelle or cr.document_key or ""
        ws.cell(row=row_idx, column=7).value = cr.response_date or ""
        ws.cell(row=row_idx, column=8).value = cr.mission or ""

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as _tmp:
        _tmp_xlsx = _tmp.name
    wb.save(_tmp_xlsx)
    shutil.copy2(_tmp_xlsx, str(output_path))
    os.remove(_tmp_xlsx)
    logger.info(
        "Orphan GED export written: %s (%d docs — MOEX En attente only, from %d total orphans)",
        output_path, len(filtered), len(set((cr.numero, cr.indice) for cr in orphan_records)),
    )


def export_orphan_summary(
    orphan_records: list[CanonicalResponse],
    all_ged_records: list[CanonicalResponse],
    output_path: Path,
) -> None:
    """
    Export a concise summary of orphan GED docs (MOEX En attente only):
    NUMERO | EMETTEUR | DATE DE RECEPTION (date_depot)
    One row per unique (NUMERO, INDICE). Saves via /tmp.
    """
    from collections import defaultdict
    from processing.actors import MOEX_MISSIONS

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Build (NUMERO, INDICE) → all responses, to check MOEX status + get date_depot
    doc_responses: dict[tuple, list[CanonicalResponse]] = defaultdict(list)
    for cr in all_ged_records:
        doc_responses[(cr.numero, cr.indice)].append(cr)

    filtered: list[CanonicalResponse] = []
    seen: set[tuple] = set()
    for cr in orphan_records:
        key = (cr.numero or "", cr.indice or "")
        if key in seen:
            continue
        responses = doc_responses.get((cr.numero, cr.indice), [])
        moex_responses = [r for r in responses if r.mission in MOEX_MISSIONS]
        if not moex_responses:
            continue
        moex_statuses = [r.raw_status.strip().lower() for r in moex_responses if r.raw_status]
        if not moex_statuses or not all(s == "en attente" for s in moex_statuses):
            continue
        seen.add(key)
        filtered.append(cr)

    filtered.sort(key=lambda r: (r.emetteur or "", r.numero or ""))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orphan Summary"

    headers = ["NUMERO", "EMETTEUR", "DATE DE RECEPTION"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = h
        cell.font = Font(bold=True)

    for row_idx, cr in enumerate(filtered, start=2):
        ws.cell(row=row_idx, column=1).value = cr.numero or ""
        ws.cell(row=row_idx, column=2).value = cr.emetteur or ""
        ws.cell(row=row_idx, column=3).value = cr.date_depot or ""

    # Auto-width columns
    for col_idx, _ in enumerate(headers, start=1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, len(filtered) + 2)),
            default=10,
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as _tmp:
        _tmp_xlsx = _tmp.name
    wb.save(_tmp_xlsx)
    shutil.copy2(_tmp_xlsx, str(output_path))
    os.remove(_tmp_xlsx)
    logger.info("Orphan summary written: %s (%d rows)", output_path, len(filtered))


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------

def export_evidence_csv(evidence: list[SourceEvidence], path: Path) -> None:
    """Write evidence_export.csv via temp file + shutil.copy2."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False,
                                     encoding="utf-8", newline="") as tmp:
        tmp_path = tmp.name
        writer = csv.DictWriter(tmp, fieldnames=[
            "sheet_name", "gf_row", "column_name", "old_value", "new_value",
            "source_type", "source_file", "source_row", "update_reason"
        ])
        writer.writeheader()
        for ev in evidence:
            writer.writerow(ev.to_dict())
    shutil.copy2(tmp_path, str(path))
    os.remove(tmp_path)
    logger.info("Evidence export written: %s (%d rows)", path, len(evidence))


def export_match_summary_csv(summary_rows: list[dict], path: Path) -> None:
    """Write match_summary.csv via temp file + shutil.copy2."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False,
                                     encoding="utf-8", newline="") as tmp:
        tmp_path = tmp.name
        writer = csv.DictWriter(tmp, fieldnames=["match_level", "count", "percentage"])
        writer.writeheader()
        writer.writerows(summary_rows)
    shutil.copy2(tmp_path, str(path))
    os.remove(tmp_path)
    logger.info("Match summary written: %s", path)
