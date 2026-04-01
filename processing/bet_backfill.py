"""
bet_backfill.py — BET Report Backfill into GrandFichier LOT sheets
JANSA GrandFichier Updater — Step 8

Reads the 4 RAPPORT_* sheets already written into the GrandFichier workbook
by bet_gf_writer.py, then back-fills the corresponding approbateur columns
in each LOT sheet.

Rules:
- GF-master driven: iterate GF LOT rows, look up matching BET records
- OLD sheets: skipped for writing (same rule as matcher.py)
- Column fill: skip if current_statut already non-empty (GED already wrote it)
- Terrell: OBS-ONLY — never writes STATUT/DATE/N° columns, only OBSERVATIONS
- OBSERVATIONS: append-only, dedup by group (existing logic) or by content (Terrell)
- History: RAPPORT_* sheets are preserved in the output workbook so future runs
  detect already-treated reports via upsert keys in bet_gf_writer.py
"""

import logging
from pathlib import Path
from typing import Optional

import openpyxl

from processing.canonical import normalize_numero
from processing.config import (
    GF_DATA_START_ROW,
    GF_APPROBATEUR_ROW,
    OLD_SHEET_PREFIX,
    BET_CONSULTANT_APPRO_NAMES,
    BET_OBS_ONLY_CONSULTANTS,
    BET_CONSULTANT_GROUP,
    BET_CONSULTANT_SHEET,
    BET_FIELD_STATUT,
    BET_FIELD_DATE,
    BET_FIELD_DATE_ALT,
    BET_FIELD_RAPPORT_ID,
    BET_FIELD_COMMENTAIRE,
    BET_FIELD_OBSERVATIONS,
    BET_FIELD_NUMERO,
    BET_FIELD_INDICE,
)
from processing.models import GFRow, GFApprobateur, SourceEvidence, AnomalyRecord
from processing.anomalies import AnomalyLogger
from processing.obs_helpers import (
    _GROUP_DISPLAY_NAMES,
    _OBS_GROUP_NORMALIZE_MAP,
    _normalize_obs_group,
    _detect_existing_obs_groups,
    _build_obs_entry,
    _is_empty_comment,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# BETReportIndex
# ---------------------------------------------------------------------------

class BETReportIndex:
    """
    Reads the 4 RAPPORT_* sheets from an already-written GrandFichier workbook
    and builds a lookup index: (normalized_numero, indice_upper) → list[dict]

    Each dict is a flat record with all column values from the RAPPORT_* sheet.
    """

    def __init__(self, gf_path: str | Path):
        self._index: dict[tuple[str, str], list[dict]] = {}
        self._load(Path(gf_path))

    def _load(self, gf_path: Path) -> None:
        """Load all 4 RAPPORT_* sheets from the workbook into the index."""
        wb = openpyxl.load_workbook(str(gf_path), data_only=True)
        total = 0

        for consultant_key, sheet_name in BET_CONSULTANT_SHEET.items():
            if sheet_name not in wb.sheetnames:
                logger.warning(
                    "BETReportIndex: sheet '%s' not found in workbook — "
                    "run run_bet_ingest.py first", sheet_name
                )
                continue

            ws = wb[sheet_name]
            if ws.max_row < 2:
                logger.info("BETReportIndex: sheet '%s' is empty", sheet_name)
                continue

            # Read headers from row 1
            headers = []
            for col in range(1, (ws.max_column or 20) + 1):
                val = ws.cell(row=1, column=col).value
                headers.append(str(val).strip() if val else '')

            # Read data rows
            sheet_count = 0
            for row_num in range(2, ws.max_row + 1):
                row_dict: dict = {
                    '_consultant': consultant_key,
                    '_sheet': sheet_name,
                    '_row': row_num,
                }
                for col_idx, header in enumerate(headers, 1):
                    if header:
                        val = ws.cell(row=row_num, column=col_idx).value
                        row_dict[header] = str(val).strip() if val is not None else ''

                # Extract NUMERO and INDICE
                numero_raw = row_dict.get(BET_FIELD_NUMERO, '')
                indice_raw = row_dict.get(BET_FIELD_INDICE, '')

                if not numero_raw:
                    continue  # Skip rows without NUMERO

                num_norm = normalize_numero(numero_raw)
                indice_norm = indice_raw.strip().upper() if indice_raw else ''

                if num_norm:
                    key = (num_norm, indice_norm)
                    self._index.setdefault(key, []).append(row_dict)
                    sheet_count += 1

            total += sheet_count
            logger.info(
                "BETReportIndex: loaded %d records from sheet '%s'",
                sheet_count, sheet_name
            )

        wb.close()
        logger.info(
            "BETReportIndex: total %d records indexed across all RAPPORT_* sheets", total
        )

    def find(self, numero: str, indice: str) -> list[dict]:
        """
        Find all BET records matching (NUMERO, INDICE).
        Returns list of record dicts, possibly from multiple consultants.
        Guard: if INDICE is empty in the BET record but not in GFRow,
        still return results and log a warning in the caller.
        """
        num_norm = normalize_numero(numero)
        indice_norm = indice.strip().upper() if indice else ''
        if not num_norm:
            return []

        # Primary: exact (numero, indice) match
        results = self._index.get((num_norm, indice_norm), [])

        # Fallback: if GFRow has an INDICE but BET records have empty INDICE,
        # return those too (INDICE extraction from PDF may be incomplete)
        if not results and indice_norm:
            fallback = self._index.get((num_norm, ''), [])
            if fallback:
                logger.debug(
                    "BETReportIndex.find: NUMERO '%s' found with empty INDICE in BET "
                    "(GFRow INDICE='%s') — using fallback empty-indice records",
                    numero, indice_norm
                )
            results = fallback

        return results

    def find_by_numero_only(self, numero: str) -> list[dict]:
        """
        Fallback: find all BET records matching NUMERO regardless of INDICE.
        Used only for diagnostic logging — NOT for writing decisions.
        """
        num_norm = normalize_numero(numero)
        if not num_norm:
            return []
        results = []
        for (n, _), records in self._index.items():
            if n == num_norm:
                results.extend(records)
        return results

    @property
    def total_records(self) -> int:
        return sum(len(v) for v in self._index.values())


# ---------------------------------------------------------------------------
# Helpers — column resolution
# ---------------------------------------------------------------------------

def _find_appro_col(gf_row: GFRow, consultant_key: str) -> Optional[GFApprobateur]:
    """
    Find the GFApprobateur object for a given consultant in a GF row.
    Matching is case-insensitive substring / exact match against Row 8 names.
    Returns None if this sheet has no column for this consultant.
    """
    candidate_names = BET_CONSULTANT_APPRO_NAMES.get(consultant_key, [])
    candidate_names_upper = [n.upper() for n in candidate_names]

    for appro in gf_row.approbateurs:
        appro_name_upper = appro.name.strip().upper()
        for candidate in candidate_names_upper:
            if appro_name_upper == candidate or candidate in appro_name_upper:
                return appro

    return None


# ---------------------------------------------------------------------------
# Helpers — date / comment extraction
# ---------------------------------------------------------------------------

def _get_bet_date(record: dict) -> str:
    """
    Extract the report date from a BET record.
    Terrell uses DATE_RECEPT; others use DATE_FICHE.
    """
    date = record.get(BET_FIELD_DATE, '').strip()
    if not date:
        date = record.get(BET_FIELD_DATE_ALT, '').strip()
    return date


def _get_bet_comment(record: dict) -> str:
    """
    Extract the observation/comment text from a BET record.
    - Terrell: uses OBSERVATIONS field
    - Others: uses COMMENTAIRE field
    """
    consultant = record.get('_consultant', '')
    if consultant == 'terrell':
        return record.get(BET_FIELD_OBSERVATIONS, '').strip()
    return record.get(BET_FIELD_COMMENTAIRE, '').strip()


# ---------------------------------------------------------------------------
# Helpers — OBSERVATIONS logic
# ---------------------------------------------------------------------------

def _should_append_obs_for_group(
    existing_obs_text: str,
    consultant_key: str,
) -> bool:
    """
    Returns True if the group for this consultant is NOT yet present
    in the existing OBSERVATIONS cell content.
    Uses the _detect_existing_obs_groups dedup logic from obs_helpers.
    """
    group = BET_CONSULTANT_GROUP.get(consultant_key, '')
    existing_groups = _detect_existing_obs_groups(existing_obs_text)
    return group not in existing_groups


def _obs_already_contains_text(existing_obs_text: str, new_text: str) -> bool:
    """
    Substring check for Terrell OBS dedup.
    Returns True if new_text is already contained in existing_obs_text.
    Case-insensitive, strip whitespace.
    """
    if not new_text or not existing_obs_text:
        return False
    return new_text.strip().lower() in existing_obs_text.strip().lower()


def _build_bet_obs_entry(consultant_key: str, record: dict) -> str:
    """
    Build the OBSERVATIONS entry string for a BET record.

    For non-Terrell: uses standard format GROUP_DISPLAY : STATUT\ncomment
    For Terrell (OBS-ONLY): uses format BET STR : [rapport ref]\ncomment
      (no STATUT since STATUT comes from GED, not from the PDF report)
    """
    group = BET_CONSULTANT_GROUP.get(consultant_key, consultant_key)
    display_name = _GROUP_DISPLAY_NAMES.get(group, group)
    comment = _get_bet_comment(record)
    statut = record.get(BET_FIELD_STATUT, '').strip()
    rapport_id = record.get(BET_FIELD_RAPPORT_ID, '').strip()

    if consultant_key == 'terrell':
        # OBS-ONLY: reference the rapport source, not the statut
        header = f"{display_name} [{rapport_id}]"
        if comment and not _is_empty_comment(comment):
            return f"{header}\n{comment}"
        return header
    else:
        # Standard format: GROUP : STATUT\ncomment
        return _build_obs_entry(display_name, statut, comment)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _write_cell(ws, row: int, col: int, value: str) -> None:
    """Write a value to a cell, preserving existing cell style (font/fill)."""
    cell = ws.cell(row=row, column=col)
    cell.value = value if value else None


def _get_obs_col(ws, gf_row: GFRow) -> Optional[int]:
    """
    Find the OBSERVATIONS column index (1-indexed) for this GF row's sheet.
    Scans Row 8 after the last approbateur group.
    Returns None if not found.
    """
    if not gf_row.approbateurs:
        return None
    last_appro = gf_row.approbateurs[-1]
    # col_statut is 0-indexed → +1 for 1-indexed, then +1 to step past it
    search_start = last_appro.col_statut + 2  # 1-indexed, one past last statut
    for c in range(search_start, min(search_start + 6, (ws.max_column or 50) + 1)):
        val = str(ws.cell(row=8, column=c).value or '').strip().upper()
        if 'OBS' in val:
            return c
    return None


def _pick_best_record(records: list[dict]) -> dict:
    """
    From multiple BET records for the same (NUMERO, INDICE, consultant),
    pick the best one to use for column fill + OBS.
    Strategy: prefer the record with the most recent DATE_FICHE,
    fallback to first record.
    """
    if len(records) == 1:
        return records[0]

    def date_key(r: dict) -> str:
        d = r.get(BET_FIELD_DATE, '') or r.get(BET_FIELD_DATE_ALT, '')
        return d or ''

    dated = [r for r in records if date_key(r)]
    if dated:
        return max(dated, key=date_key)
    return records[0]


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def backfill_bet_reports(
    gf_workbook_path: str | Path,
    gf_rows: list[GFRow],
    anomaly_logger: AnomalyLogger,
    output_path: str | Path,
) -> tuple[list[SourceEvidence], int]:
    """
    Main BET backfill function.

    Args:
        gf_workbook_path : path to the GrandFichier workbook already updated by
                           Temps 1 (GED pass) AND containing the 4 RAPPORT_* sheets
                           written by bet_gf_writer.py
        gf_rows          : list[GFRow] read by grandfichier_reader (same run context)
        anomaly_logger   : shared anomaly logger (appends to existing log)
        output_path      : where to save the updated workbook (can be same as input)

    Returns:
        evidence_records : list[SourceEvidence] for all fields written
        fields_updated   : total count of fields written
    """
    gf_workbook_path = Path(gf_workbook_path)
    output_path = Path(output_path)

    logger.info("=" * 50)
    logger.info("Step 8: BET Report Backfill")
    logger.info("  Input workbook : %s", gf_workbook_path)
    logger.info("  Output path    : %s", output_path)
    logger.info("=" * 50)

    # Guard: verify workbook can be opened before proceeding
    try:
        _test_wb = openpyxl.load_workbook(str(gf_workbook_path), read_only=True)
        _test_wb.close()
    except Exception as e:
        logger.error(
            "Step 8: Cannot open workbook %s: %s — aborting BET backfill",
            gf_workbook_path, e
        )
        return [], 0

    # ── Step 8.1: Build BETReportIndex from RAPPORT_* sheets ──────────────
    logger.info("Step 8.1: Building BET report index from RAPPORT_* sheets...")
    bet_index = BETReportIndex(gf_workbook_path)
    if bet_index.total_records == 0:
        logger.warning(
            "Step 8: BETReportIndex is empty — no RAPPORT_* sheets found or all empty. "
            "Run run_bet_ingest.py first to populate them. Skipping BET backfill."
        )
        return [], 0
    logger.info("Step 8.1: Index built — %d total BET records", bet_index.total_records)

    # ── Step 8.2: Load workbook for writing ───────────────────────────────
    logger.info("Step 8.2: Loading workbook for writing...")
    wb = openpyxl.load_workbook(str(gf_workbook_path))

    evidence_records: list[SourceEvidence] = []
    fields_updated = 0

    # Counters for summary
    stats = {
        'rows_processed':       0,
        'rows_old_skipped':     0,
        'matches_found':        0,
        'columns_written':      0,
        'columns_noop_filled':  0,   # already filled by GED pass
        'obs_appended':         0,
        'obs_noop':             0,
        'no_gf_column':         0,
        'unmatched':            0,
    }

    # ── Step 8.3: Main loop — GF-master driven ────────────────────────────
    logger.info("Step 8.3: Iterating GF rows for BET backfill...")
    for gf_row in gf_rows:

        # RULE: Skip OLD sheets entirely — same as matcher.py
        if gf_row.sheet_name.startswith(OLD_SHEET_PREFIX):
            stats['rows_old_skipped'] += 1
            continue

        stats['rows_processed'] += 1

        # Look up matching BET records
        bet_records = bet_index.find(gf_row.numero, gf_row.indice)
        if not bet_records:
            stats['unmatched'] += 1
            # Diagnostic: check if NUMERO matches but INDICE differs
            by_num = bet_index.find_by_numero_only(gf_row.numero)
            if by_num:
                logger.debug(
                    "GF row %s/%s: NUMERO '%s' found in BET index but INDICE '%s' "
                    "has no match (BET indices: %s)",
                    gf_row.sheet_name, gf_row.row_number,
                    gf_row.numero, gf_row.indice,
                    list({r.get(BET_FIELD_INDICE, '') for r in by_num})
                )
            continue

        stats['matches_found'] += 1

        # Get the worksheet for this GF row
        if gf_row.sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in workbook — skipping", gf_row.sheet_name)
            continue
        ws = wb[gf_row.sheet_name]

        # Group BET records by consultant key
        records_by_consultant: dict[str, list[dict]] = {}
        for rec in bet_records:
            ckey = rec.get('_consultant', '')
            records_by_consultant.setdefault(ckey, []).append(rec)

        # Read current OBSERVATIONS cell value once
        from processing.grandfichier_reader import _cell_str
        obs_col = _get_obs_col(ws, gf_row)
        current_obs = ''
        if obs_col:
            current_obs = _cell_str(ws, gf_row.row_number, obs_col) or ''

        obs_additions: list[str] = []  # collect new OBS entries for this row

        for consultant_key, consultant_records in records_by_consultant.items():

            # Pick best record: prefer most recent DATE_FICHE, else first
            best_record = _pick_best_record(consultant_records)

            # ── COLUMN FILL (not for Terrell) ────────────────────────────
            if consultant_key not in BET_OBS_ONLY_CONSULTANTS:
                appro = _find_appro_col(gf_row, consultant_key)

                if appro is None:
                    # This LOT sheet has no column for this consultant — normal
                    stats['no_gf_column'] += 1
                    anomaly_logger.log(AnomalyRecord(
                        anomaly_type='NO_GF_COLUMN',
                        severity='INFO',
                        source_type='REPORT',
                        source_file=best_record.get(BET_FIELD_RAPPORT_ID, ''),
                        source_row_or_page=f"row {best_record.get('_row', '')}",
                        document_key=gf_row.document_key,
                        description=(
                            f"Consultant '{consultant_key}' has no approbateur column "
                            f"in sheet '{gf_row.sheet_name}'"
                        ),
                        raw_data={'consultant': consultant_key, 'sheet': gf_row.sheet_name},
                    ))
                else:
                    # Check if column is already filled (GED pass wrote it)
                    if appro.current_statut.strip():
                        # Already filled — NOOP for columns
                        stats['columns_noop_filled'] += 1
                        logger.debug(
                            "GF %s/row%d [%s]: STATUT already '%s' — column NOOP",
                            gf_row.sheet_name, gf_row.row_number,
                            consultant_key, appro.current_statut
                        )
                    else:
                        # Write DATE / RAPPORT_ID / STATUT
                        new_date   = _get_bet_date(best_record)
                        new_num    = best_record.get(BET_FIELD_RAPPORT_ID, '')
                        new_statut = best_record.get(BET_FIELD_STATUT, '')

                        _write_cell(ws, gf_row.row_number, appro.col_date + 1,   new_date)
                        _write_cell(ws, gf_row.row_number, appro.col_num + 1,    new_num)
                        _write_cell(ws, gf_row.row_number, appro.col_statut + 1, new_statut)

                        stats['columns_written'] += 1
                        fields_updated += 3

                        # Log SourceEvidence for each written field
                        for col_name, col_idx, new_val in [
                            ('DATE',   appro.col_date + 1,   new_date),
                            ('N°',     appro.col_num + 1,    new_num),
                            ('STATUT', appro.col_statut + 1, new_statut),
                        ]:
                            evidence_records.append(SourceEvidence(
                                sheet_name=gf_row.sheet_name,
                                row_number=gf_row.row_number,
                                column_name=f"{appro.name} / {col_name}",
                                old_value='',
                                new_value=str(new_val),
                                source_type='REPORT',
                                source_file=best_record.get(BET_FIELD_RAPPORT_ID, ''),
                                source_row_or_page=f"row {best_record.get('_row', '')}",
                                update_reason=f"BET backfill — {consultant_key} report",
                            ))

                        logger.debug(
                            "GF %s/row%d [%s]: wrote STATUT=%s DATE=%s N°=%s",
                            gf_row.sheet_name, gf_row.row_number,
                            consultant_key, new_statut, new_date, new_num
                        )

            # ── OBSERVATIONS (all consultants including Terrell) ─────────
            comment_text = _get_bet_comment(best_record)
            if not comment_text:
                continue  # No comment to append

            if consultant_key == 'terrell':
                # Terrell dedup: substring check on comment content
                if _obs_already_contains_text(current_obs, comment_text):
                    stats['obs_noop'] += 1
                    logger.debug(
                        "GF %s/row%d [terrell]: OBS already contains comment — NOOP",
                        gf_row.sheet_name, gf_row.row_number
                    )
                else:
                    new_entry = _build_bet_obs_entry(consultant_key, best_record)
                    obs_additions.append(new_entry)
            else:
                # Standard dedup: by consultant group name
                if not _should_append_obs_for_group(current_obs, consultant_key):
                    stats['obs_noop'] += 1
                    logger.debug(
                        "GF %s/row%d [%s]: OBS group already present — NOOP",
                        gf_row.sheet_name, gf_row.row_number, consultant_key
                    )
                else:
                    new_entry = _build_bet_obs_entry(consultant_key, best_record)
                    obs_additions.append(new_entry)

        # Write accumulated OBSERVATIONS additions for this row
        if obs_additions and obs_col:
            separator = '\n' if current_obs.strip() else ''
            new_obs_text = current_obs + separator + '\n'.join(obs_additions)
            _write_cell(ws, gf_row.row_number, obs_col, new_obs_text)

            stats['obs_appended'] += len(obs_additions)
            fields_updated += 1

            evidence_records.append(SourceEvidence(
                sheet_name=gf_row.sheet_name,
                row_number=gf_row.row_number,
                column_name='OBSERVATIONS',
                old_value=current_obs,
                new_value=new_obs_text,
                source_type='REPORT',
                source_file='; '.join(set(
                    r.get(BET_FIELD_RAPPORT_ID, '') for r in bet_records
                )),
                source_row_or_page='BET backfill',
                update_reason=f"BET backfill — appended {len(obs_additions)} new entry(ies)",
            ))

    # ── Step 8.4: Save workbook ───────────────────────────────────────────
    logger.info("Step 8.4: Saving updated workbook to %s...", output_path)
    wb.save(str(output_path))
    wb.close()
    logger.info("Step 8.4: Workbook saved.")

    # ── Step 8.5: Log summary ─────────────────────────────────────────────
    logger.info("=" * 50)
    logger.info("Step 8 BET Backfill Summary:")
    logger.info("  GF rows processed        : %d", stats['rows_processed'])
    logger.info("  GF rows skipped (OLD)    : %d", stats['rows_old_skipped'])
    logger.info("  GF rows matched to BET   : %d", stats['matches_found'])
    logger.info("  GF rows unmatched        : %d", stats['unmatched'])
    logger.info("  Columns written          : %d", stats['columns_written'])
    logger.info("  Columns NOOP (GED filled): %d", stats['columns_noop_filled'])
    logger.info("  OBS entries appended     : %d", stats['obs_appended'])
    logger.info("  OBS entries NOOP (dedup) : %d", stats['obs_noop'])
    logger.info("  No GF column (INFO)      : %d", stats['no_gf_column'])
    logger.info("  Total fields written     : %d", fields_updated)
    logger.info("=" * 50)

    return evidence_records, fields_updated
