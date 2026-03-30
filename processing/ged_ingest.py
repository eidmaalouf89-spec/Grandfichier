"""
JANSA GrandFichier Updater — GED dump ingestion (V1)

Reads the AxeoBIM GED Excel export ("Vue détaillée des documents" sheet)
and produces a list of CanonicalResponse records.

One CanonicalResponse per (document × mission response) row.
Rows with no Mission value (col 25 empty) are skipped.
"""
import logging
from pathlib import Path
from typing import Optional

import openpyxl

from processing.config import (
    GED_SHEET_PRIMARY, GED_SHEET_VARIANT,
    GED_DATA_START_ROW, GED_COL,
)
from processing.models import CanonicalResponse
from processing.dates import parse_date, date_to_str, parse_delay
from processing.statuses import get_normalized_code
from processing.canonical import normalize_numero, normalize_key, _s

logger = logging.getLogger(__name__)


def _cell(row_cells, col_idx: int):
    """Safely get a cell value by 0-based column index from an openpyxl row."""
    try:
        return row_cells[col_idx].value
    except (IndexError, AttributeError):
        return None


def _str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _pick_sheet(wb: openpyxl.Workbook) -> tuple:
    """Return (worksheet, sheet_name). Prefer PRIMARY, fall back to VARIANT."""
    if GED_SHEET_PRIMARY in wb.sheetnames:
        return wb[GED_SHEET_PRIMARY], GED_SHEET_PRIMARY
    if GED_SHEET_VARIANT in wb.sheetnames:
        return wb[GED_SHEET_VARIANT], GED_SHEET_VARIANT
    raise ValueError(
        f"GED workbook contains neither '{GED_SHEET_PRIMARY}' nor '{GED_SHEET_VARIANT}'. "
        f"Available sheets: {wb.sheetnames}"
    )


def ingest_ged(
    excel_path: str | Path,
    status_map: dict,
) -> tuple[list[CanonicalResponse], list[dict]]:
    """
    Parse a GED Excel dump and return:
      - list[CanonicalResponse]: one record per (document × mission) row with a mission value
      - list[dict]: skipped rows (for diagnostics)

    Args:
        excel_path: path to the GED .xlsx file
        status_map: loaded status_map.json dict (for normalization)
    """
    excel_path = Path(excel_path)
    source_file = excel_path.name
    records: list[CanonicalResponse] = []
    skipped: list[dict] = []

    logger.info("Loading GED file: %s", excel_path)
    wb = openpyxl.load_workbook(str(excel_path), data_only=True, read_only=True)

    ws, sheet_name = _pick_sheet(wb)
    logger.info("Using sheet: '%s'", sheet_name)

    row_idx = 0
    skipped_no_mission = 0
    skipped_empty = 0

    for row in ws.iter_rows(min_row=GED_DATA_START_ROW):
        row_idx += 1
        row_cells = list(row)

        # Skip completely empty rows
        if all(c.value is None for c in row_cells):
            skipped_empty += 1
            continue

        # Extract all fields
        affaire    = _str(_cell(row_cells, GED_COL["affaire"]))
        projet     = _str(_cell(row_cells, GED_COL["projet"]))
        batiment   = _str(_cell(row_cells, GED_COL["batiment"]))
        phase      = _str(_cell(row_cells, GED_COL["phase"]))
        emetteur   = _str(_cell(row_cells, GED_COL["emetteur"]))
        specialite = _str(_cell(row_cells, GED_COL["specialite"]))
        lot        = _str(_cell(row_cells, GED_COL["lot"]))
        type_doc   = _str(_cell(row_cells, GED_COL["type_doc"]))
        zone       = _str(_cell(row_cells, GED_COL["zone"]))
        niveau     = _str(_cell(row_cells, GED_COL["niveau"]))
        numero_raw = _cell(row_cells, GED_COL["numero"])
        indice     = _str(_cell(row_cells, GED_COL["indice"]))
        libelle    = _str(_cell(row_cells, GED_COL["libelle"]))

        mission      = _str(_cell(row_cells, GED_COL["mission"]))
        respondant   = _str(_cell(row_cells, GED_COL["respondant"]))
        date_limite  = _cell(row_cells, GED_COL["date_limite"])
        reponse_le   = _cell(row_cells, GED_COL["reponse_le"])
        ecart        = _cell(row_cells, GED_COL["ecart_reponse"])
        reponse_raw  = _str(_cell(row_cells, GED_COL["reponse"]))
        commentaire  = _str(_cell(row_cells, GED_COL["commentaire"]))
        pieces       = _str(_cell(row_cells, GED_COL["pieces_jointes"]))

        # Skip rows with no mission (document metadata rows without reviewer)
        if not mission:
            skipped_no_mission += 1
            continue

        # Normalize numero
        numero = normalize_numero(str(numero_raw) if numero_raw is not None else "")

        # Build display key for traceability (V3.0: not used for matching)
        doc_key = normalize_key(
            f"{_s(lot)}/{_s(type_doc)}/{_s(str(numero_raw) if numero_raw is not None else '')}/{_s(indice)}"
        )

        # Normalize status
        norm_status = get_normalized_code(reponse_raw, status_map)

        # Parse dates
        response_date_obj = parse_date(reponse_le)
        deadline_date_obj = parse_date(date_limite)
        days_delta_val = parse_delay(ecart)

        parse_warnings = []
        if not affaire and not projet:
            parse_warnings.append("Both AFFAIRE and PROJET are empty")
        if not numero:
            parse_warnings.append("NUMERO is empty or unparseable")

        source_row = f"row {GED_DATA_START_ROW + row_idx - 1}"

        rec = CanonicalResponse(
            source_type="GED",
            source_file=source_file,
            source_row_or_page=source_row,
            document_key=doc_key,
            lot=lot,
            type_doc=type_doc,
            numero=numero,
            indice=indice,
            batiment=batiment,
            zone=zone,
            niveau=niveau,
            emetteur=emetteur,
            mission=mission,
            respondant=respondant,
            raw_status=reponse_raw,
            normalized_status=norm_status,
            response_date=date_to_str(response_date_obj),
            deadline_date=date_to_str(deadline_date_obj),
            days_delta=days_delta_val,
            comment=commentaire,
            attachments=pieces,
            parse_warnings=parse_warnings,
        )
        records.append(rec)

    wb.close()

    logger.info(
        "GED ingest complete: %d records ingested, %d skipped (no mission), %d empty rows skipped",
        len(records), skipped_no_mission, skipped_empty,
    )
    return records, skipped
