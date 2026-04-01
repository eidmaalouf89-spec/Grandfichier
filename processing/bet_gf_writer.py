"""
bet_gf_writer.py — BET Report GrandFichier Sheet Writer
JANSA VISASIST — BET PDF Report Ingestion
Version 1.0 — April 2026

Writes parser output records from 4 BET consultants into dedicated sheets
in the GrandFichier Excel workbook using openpyxl.

Public API:
    write_bet_reports_to_gf(gf_path, records_by_consultant) -> stats
"""

import logging
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Sheet names (exact, case-sensitive)
# ---------------------------------------------------------------------------

SHEET_NAMES = {
    'lesommer': 'RAPPORT_LE_SOMMER',
    'avls':     'RAPPORT_AVLS',
    'terrell':  'RAPPORT_TERRELL',
    'socotec':  'RAPPORT_SOCOTEC',
}

# ---------------------------------------------------------------------------
# Column definitions per sheet
# ---------------------------------------------------------------------------

# Common columns A–K (all sheets)
COMMON_COLUMNS = [
    'SOURCE', 'RAPPORT_ID', 'DATE_FICHE', 'NUMERO', 'INDICE',
    'REF_DOC', 'STATUT_NORM', 'COMMENTAIRE', 'PDF_PAGE',
    'UPSERT_KEY', 'LAST_UPDATED',
]

# Extra columns per sheet (appended after col K)
EXTRA_COLUMNS = {
    'RAPPORT_LE_SOMMER': ['LOT_TYPE', 'SECTION', 'TABLE_TYPE'],
    'RAPPORT_AVLS':      ['LOT_LABEL', 'LOT_NUM', 'N_VISA', 'REVIEWER'],
    'RAPPORT_TERRELL':   ['BAT', 'LOT', 'SPECIALITE', 'TYPE_DOC', 'NIVEAU',
                          'DATE_SOURCE', 'DESIGNATION'],
    'RAPPORT_SOCOTEC':   ['CT_REF', 'OBS_NUM'],
}

# Map from record field names to the common column headers
# (some parsers use different field names)
FIELD_MAP_COMMON = {
    'SOURCE': 'SOURCE',
    'RAPPORT_ID': 'RAPPORT_ID',
    'DATE_FICHE': 'DATE_FICHE',       # lesommer uses DATE_VISA, avls/socotec use DATE_FICHE
    'NUMERO': 'NUMERO',
    'INDICE': 'INDICE',
    'REF_DOC': 'REF_DOC',
    'STATUT_NORM': 'STATUT_NORM',
    'COMMENTAIRE': 'COMMENTAIRE',     # lesommer: COMMENTAIRE, terrell: OBSERVATIONS
    'PDF_PAGE': 'PDF_PAGE',
}

# ---------------------------------------------------------------------------
# Upsert key functions per sheet
# ---------------------------------------------------------------------------

UPSERT_KEYS = {
    'RAPPORT_LE_SOMMER': lambda r: f"{r.get('RAPPORT_ID','')}|{r.get('NUMERO','')}|{r.get('INDICE','')}|{r.get('SECTION','')}|{r.get('TABLE_TYPE','')}",
    'RAPPORT_AVLS':      lambda r: f"{r.get('RAPPORT_ID','')}|{r.get('REF_DOC','')}|{r.get('LOT_LABEL','')}|{r.get('N_VISA','')}",
    'RAPPORT_TERRELL':   lambda r: f"{r.get('RAPPORT_ID','')}|{r.get('NUMERO','')}|{r.get('INDICE','')}|{r.get('BAT','')}|{r.get('LOT','')}",
    'RAPPORT_SOCOTEC':   lambda r: f"{r.get('RAPPORT_ID','')}|{r.get('NUMERO','')}|{r.get('STATUT_NORM','')}",
}

# ---------------------------------------------------------------------------
# Header style
# ---------------------------------------------------------------------------

JANSA_TEAL = 'FF006B6B'
HEADER_FILL = PatternFill(start_color=JANSA_TEAL, end_color=JANSA_TEAL, fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFFFF')


def _apply_header_style(ws, ncols: int) -> None:
    """Apply JANSA teal fill + white bold font to row 1."""
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _auto_width(ws, max_width: int = 60) -> None:
    """Set column widths based on max content length, capped at max_width."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                if val_len > max_len:
                    max_len = val_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


# ---------------------------------------------------------------------------
# Record field extraction helpers
# ---------------------------------------------------------------------------

def _get_date_fiche(record: dict) -> str:
    """Get the report date, checking multiple possible field names."""
    return (record.get('DATE_FICHE')
            or record.get('DATE_VISA')
            or record.get('DATE_RECEPT')
            or '')


def _get_commentaire(record: dict) -> str:
    """Get observations/comments, checking multiple possible field names."""
    raw = (record.get('COMMENTAIRE')
           or record.get('OBSERVATIONS')
           or '')
    return str(raw)[:200] if raw else ''


def _build_row_values(record: dict, all_columns: list, key_fn, timestamp: str) -> list:
    """Build a flat list of cell values for one record row."""
    values = []
    for col in all_columns:
        if col == 'SOURCE':
            values.append(record.get('SOURCE', ''))
        elif col == 'RAPPORT_ID':
            values.append(record.get('RAPPORT_ID', ''))
        elif col == 'DATE_FICHE':
            values.append(_get_date_fiche(record))
        elif col == 'NUMERO':
            values.append(record.get('NUMERO', ''))
        elif col == 'INDICE':
            values.append(record.get('INDICE', ''))
        elif col == 'REF_DOC':
            values.append(record.get('REF_DOC', ''))
        elif col == 'STATUT_NORM':
            values.append(record.get('STATUT_NORM', ''))
        elif col == 'COMMENTAIRE':
            values.append(_get_commentaire(record))
        elif col == 'PDF_PAGE':
            values.append(record.get('PDF_PAGE', ''))
        elif col == 'UPSERT_KEY':
            values.append(key_fn(record))
        elif col == 'LAST_UPDATED':
            values.append(timestamp)
        # Extra columns — direct field lookup
        elif col == 'CT_REF':
            # SOCOTEC stores CT ref in RAPPORT_ID
            values.append(record.get('RAPPORT_ID', record.get('CT_REF', '')))
        else:
            values.append(record.get(col, ''))
    return values


# ---------------------------------------------------------------------------
# Upsert logic
# ---------------------------------------------------------------------------

def upsert_sheet(ws, records: list[dict], sheet_key: str, key_fn,
                 all_columns: list) -> dict:
    """
    Upsert records into an openpyxl worksheet.

    Rules:
    - Row 1 = header (never overwrite)
    - Build index: {upsert_key → row_number} from existing data
    - For each record:
        key = key_fn(record)
        if key not in index → APPEND new row
        elif STATUT_NORM changed → UPDATE that row (cols G, H, K only)
        else → NO-OP

    Returns stats: {"inserted": int, "updated": int, "noop": int}
    """
    timestamp = datetime.now(timezone.utc).isoformat()
    stats = {'inserted': 0, 'updated': 0, 'noop': 0}

    # Find column indices (1-based) for key columns in worksheet
    key_col_idx    = all_columns.index('UPSERT_KEY') + 1   # col J (1-based)
    statut_col_idx = all_columns.index('STATUT_NORM') + 1   # col G
    comment_col_idx = all_columns.index('COMMENTAIRE') + 1  # col H
    ts_col_idx     = all_columns.index('LAST_UPDATED') + 1  # col K

    # Build existing key → row index
    existing_keys: dict[str, int] = {}
    for row_num in range(2, ws.max_row + 1):
        key_val = ws.cell(row=row_num, column=key_col_idx).value
        if key_val:
            existing_keys[str(key_val)] = row_num

    for record in records:
        upsert_key = key_fn(record)

        if upsert_key not in existing_keys:
            # INSERT: append new row
            row_values = _build_row_values(record, all_columns, key_fn, timestamp)
            ws.append(row_values)
            existing_keys[upsert_key] = ws.max_row
            stats['inserted'] += 1

        else:
            # Key exists — check if STATUT_NORM changed
            existing_row = existing_keys[upsert_key]
            existing_statut = ws.cell(row=existing_row, column=statut_col_idx).value

            if existing_statut != record.get('STATUT_NORM'):
                # UPDATE: overwrite STATUT_NORM, COMMENTAIRE, LAST_UPDATED
                ws.cell(row=existing_row, column=statut_col_idx).value = record.get('STATUT_NORM')
                ws.cell(row=existing_row, column=comment_col_idx).value = _get_commentaire(record)
                ws.cell(row=existing_row, column=ts_col_idx).value = timestamp
                stats['updated'] += 1
            else:
                stats['noop'] += 1

    return stats


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_bet_reports_to_gf(
    gf_path: str | Path,
    records_by_consultant: dict[str, list[dict]],
) -> dict:
    """
    Write BET parser records into the GrandFichier workbook.

    Args:
        gf_path: path to the GrandFichier .xlsx file (read and written in place)
        records_by_consultant: {
            'lesommer': list[dict],
            'avls':     list[dict],
            'terrell':  list[dict],
            'socotec':  list[dict],
        }

    Returns:
        stats: {
            'RAPPORT_LE_SOMMER': {'inserted': int, 'updated': int, 'noop': int},
            'RAPPORT_AVLS':      {...},
            'RAPPORT_TERRELL':   {...},
            'RAPPORT_SOCOTEC':   {...},
        }
    """
    gf_path = Path(gf_path)
    logger.info("Opening GrandFichier: %s", gf_path)

    # Load workbook (never read_only — we need to write)
    wb = openpyxl.load_workbook(str(gf_path))

    all_stats: dict[str, dict] = {}

    for consultant_key, sheet_name in SHEET_NAMES.items():
        records = records_by_consultant.get(consultant_key, [])
        if not records:
            logger.info("No records for %s — skipping sheet %s", consultant_key, sheet_name)
            all_stats[sheet_name] = {'inserted': 0, 'updated': 0, 'noop': 0}
            continue

        logger.info("Writing %d records to sheet %s", len(records), sheet_name)

        # Get or create the sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        # Build full column list for this sheet
        extra_cols = EXTRA_COLUMNS.get(sheet_name, [])
        all_columns = COMMON_COLUMNS + extra_cols

        # Write header row if sheet is new or empty
        if ws.max_row < 1 or (ws.max_row == 1 and ws.cell(1, 1).value is None):
            for col_idx, col_name in enumerate(all_columns, 1):
                ws.cell(row=1, column=col_idx).value = col_name
            _apply_header_style(ws, len(all_columns))
            ws.freeze_panes = 'A2'

        key_fn = UPSERT_KEYS[sheet_name]
        stats = upsert_sheet(ws, records, sheet_name, key_fn, all_columns)
        all_stats[sheet_name] = stats

        # Auto-width columns
        _auto_width(ws)

        logger.info(
            "Sheet %s: inserted=%d updated=%d noop=%d",
            sheet_name, stats['inserted'], stats['updated'], stats['noop']
        )

    # Save workbook in place
    wb.save(str(gf_path))
    logger.info("GrandFichier saved: %s", gf_path)

    return all_stats
